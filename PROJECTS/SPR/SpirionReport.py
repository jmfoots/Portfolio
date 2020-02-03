#Imports
import csv, openpyxl, time, os, shutil
from datetime import date
from time import strptime, strftime
from copy import copy, deepcopy
from openpyxl import load_workbook, styles
from openpyxl.worksheet.table import Table, TableStyleInfo
import hashlib
#Variables
Date = input('Enter Date (m-d-y):\n')
Report = Date + '-Report.xlsx'
ReportSheet = 'PII Report'
ChangeSheet = 'Change Report'
TableName = 'Report'
TagSheet = 'Tags'
TagFolders = 'Folders'
Locations = 'ALL.csv'
OutOfDate = 'OutOfDate.csv'
dirPath = os.path.dirname(__file__)
filePath = os.path.join(dirPath, 'Endpoints\\')
Program = os.path.join(dirPath, 'Program\\')
Template = Program + 'Template.xlsx'
filters = Program + 'Filters.csv'
Tags = Program + 'Tags.xlsx'
#Messages
updateEndpointTrue = ' PII changed: Updated'
getTagError = 'Endpoint not found!'
getPIICountError = 'Attempted to get Count: Failed'
Failure = 'Save Failed: Workbook already open'
Success = 'Save Succeeded: Workbook upddated'
#Filters
dataTypes = ["Social Security Number", "Bank Account Number", "Credit Card Number"]
Extensions = [".pdf",".doc",".xls",".db",".png",".jp",".csv",".bmp",".txt",".ppt",".htm",".wpd",".wps",".mdb"]
#Functions
def getFilters():
    filterby = [] 
    with open(filters) as data:
        entries = list(csv.reader(data))
        for entry in entries:
            filterby.append(entry[0])
    return filterby
FalsePositives = getFilters()
def Filter(endpoint, datatype, location):
    match = False
    for ext in Extensions:
        if ext in location:
            match = True
    if match == False:
        return False
    for word in FalsePositives:
        if word.lower() in location.lower():
            return False
    return True
def updateFiles(Tag, Endpoint, entries):
    wb = openpyxl.load_workbook(Template)
    ws = wb.active
    mR = ws.max_row+1
    tagBook = openpyxl.load_workbook(Tags)
    tagFolders = tagBook[TagFolders]
    mR2 = tagFolders.max_row+1
    folderName = ''
    savedLocations = []
    for row2 in range(1, mR2):
        if Tag == tagFolders.cell(row=row2, column=1).value:
            folderName = tagFolders.cell(row=row2, column=2).value
    if folderName != '':
        style = TableStyleInfo(name="TableStyleLight16", showRowStripes=True)
        for entry in entries:
            fields = ','.join(entry).split(",")
            shift=0
            if len(fields) > 10: shift=(len(fields)-10)
            endpoint = fields[(2+shift)]
            if len(fields) > 10: location = fields[0:shift]
            else: location = fields[0]
            dataType = fields[-(4+shift)]
            if Endpoint == endpoint:
                if Filter(endpoint, dataType, location) == True and dataType in dataTypes:
                    ws.cell(row=mR, column=1).value = endpoint
                    ws.cell(row=mR, column=2).value = dataType
                    ws.cell(row=mR, column=3).value = location
                    mR+=1
                    if filePath+folderName+'\\'+Endpoint+'.xlsx' not in savedLocations:
                        savedLocations.append(filePath+folderName+'\\'+Endpoint+'.xlsx')
                    TableID = [i for i, table in enumerate(ws._tables) if table.name == TableName][0]
                    ws._tables[TableID] = Table(displayName=TableName, ref = "A1:C"+str(mR-1), tableStyleInfo=style)
                    wb.save(filePath+folderName+'\\'+Endpoint+'.xlsx')
    return savedLocations
def updateEndpoint(Endpoint, PII):
    wb = openpyxl.load_workbook(Report)
    ws = wb[ReportSheet]
    mR = ws.max_row+1
    Updates = []
    for row in range(1, mR):
        if Endpoint == ws.cell(row=row, column=2).value:
            ws.cell(row=row, column=4).value = PII[0]
            ws.cell(row=row, column=5).value = PII[1]
            ws.cell(row=row, column=6).value = PII[2]
        if Endpoint not in Updates:
            Updates.append(Endpoint)
    wb.save(Report)
    return Updates
def removeDuplicates(entries):
    read = shutil.copy(entries, "copy" + entries)
    with open(read, "r") as f_in, \
            open(entries, "w") as f_out:
        seen = set()
        for line in f_in:
            ID = ','.join(line.split(',')[0:3])+'\n'
            line_hash = hashlib.md5(ID.encode()).digest()
            if line_hash not in seen:
                seen.add(line_hash)
                f_out.write(line)
    os.remove(read)
    print(str(len(seen)) + " files")
    with open(Locations, encoding="utf-8") as data:
        return list(csv.reader(data))
def getTag(endpoint):
    wb = openpyxl.load_workbook(Tags)
    ws = wb[TagSheet]
    mR = ws.max_row+1
    for row in range(1, mR):
        if endpoint == ws.cell(row=row, column=2).value:
            return ws.cell(row=row, column=1).value
    print(endpoint + ": " + getTagError)
    return getTagError
def getPIICount(endpoint, entries):
    abbr = ["Social Security Number", "Credit Card Number", "Bank Account Number"]
    PII = [0,0,0]
    for entry in entries:
        fields = ','.join(entry).split(",")
        shift=0
        if len(fields) > 10: shift=(len(fields)-10)
        if fields[(2+shift)] == endpoint:
            Endpoint = fields[(2+shift)]
            DataType = fields[-(4+shift)]
            if len(fields) > 10: location = fields[0:shift]
            else: location = fields[0]
            if Filter(Endpoint, DataType, location) == True and DataType in dataTypes:
                try:
                    PII[abbr.index(DataType)] = PII[abbr.index(DataType)] + 1
                except:
                    print(getPIICountError)
    return PII
def resetPII():
    wb = openpyxl.load_workbook(Report)
    ws = wb[ReportSheet]
    mR = ws.max_row+1
    for row in range(2,mR-1):
        ws.cell(row=row, column=4).value = 0
        ws.cell(row=row, column=5).value = 0
        ws.cell(row=row, column=6).value = 0
    wb.save(Report)
def resetEndpoints():
    endpoints = os.path.abspath('./Endpoints')
    i=0
    for folder in os.listdir(endpoints):
        folder = os.path.abspath(os.path.join(endpoints,folder))
        for file in os.listdir(folder):
            file = os.path.join(folder, file)
            if file.endswith('.xlsx'):
                os.remove(file)
                i=i+1
    print('Removed ' + str(i) + ' Files')
def excludeOutDated():
    exclude = []
    with open(OutOfDate) as data:
        entries = list(csv.reader(data))
        for entry in entries[1:]:
            exclude.append(entry[1])
    return exclude
#Main
entries = removeDuplicates(Locations)
Endpoints=[]
resetPII()
resetEndpoints()
Outdated = excludeOutDated()
i=0
print("Updating Files\n...")
for entry in entries:
    fields = ','.join(entry).split(",")
    shift=0
    if len(fields) > 10: shift=(len(fields)-10)
    Endpoint = fields[(2+shift)]
    if Endpoint not in Endpoints and Endpoint not in Outdated:
        Endpoints.append(Endpoint)
        Tag = getTag(Endpoint)
        if Tag != getTagError:
            PII = getPIICount(Endpoint, entries)
            totalPII = str(PII[0]+PII[1]+PII[2])
            if totalPII != "0":
                print(Endpoint + ": " + totalPII)
            Updates = updateEndpoint(Endpoint, PII)
            savedLocations = updateFiles(Tag, Endpoint, entries)
            for item in savedLocations:
                i=i+1
print("Updated " + str(i) + " Files")
time.sleep(2)
        
        
        
